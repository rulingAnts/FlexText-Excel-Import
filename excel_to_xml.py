#!/usr/bin/env python3
import argparse
import os
import sys
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom 

def convert_excel_to_xml_dom(excel_path):
    """
    Core function to read interlinear data from an Excel file, validate it, 
    and return an XML Element (DOM object). Includes tqdm for progress reporting
    and an early exit for long stretches of blank rows.

    Args:
        excel_path (str): The full path to the Excel file.

    Returns:
        tuple: (xml.etree.ElementTree.Element, list) 
               The root XML element and a list of errors.
               Returns (None, list) on a fatal error.
    """
    try:
        # Import necessary external libraries
        import openpyxl 
        from tqdm import tqdm # Import tqdm here
    except ImportError as e:
        # Provide a helpful error if either library is missing
        library_name = 'openpyxl' if 'openpyxl' in str(e) else 'tqdm'
        return None, [f"FATAL ERROR: The '{library_name}' library is required. Please install it with: pip install {library_name}"]
        
    # --- Excel Template Constants (INSIDE THE FUNCTION) ---
    METADATA_CELLS = {
        'title': 'C2',
        'author': 'C3',
        'transcriber': 'C4',
        'writing_system_vernacular': 'N2',
        'writing_system_free': 'N3',
        'writing_system_gloss': 'N4'
    }
    DATA_START_ROW = 6
    DATA_START_COLUMN = 3 # Column C
    DATA_END_COLUMN = 26 # Column Z (where the free translation merge ends)
    ROWS_PER_LINE_BLOCK = 4
    BLANK_BLOCK_EXIT_THRESHOLD = 5 # Exit after 5 consecutive empty 4-row blocks (20 blank rows)
    # -----------------------------------------------------------

    error_list = []
    consecutive_empty_blocks = 0

    # --- Helper Functions (Nested for clean encapsulation) ---

    def get_cell_value(sheet, row, col):
        """Safely retrieves the value from a cell and cleans it (e.g., handles merged cells)."""
        cell = sheet.cell(row=row, column=col)
        value = cell.value

        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                if value is None:
                    value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
                break
                
        return str(value).strip() if value is not None else None

    def is_row_empty(sheet, row):
        """Checks if all cells in the data range (Col C to Z) for a row are empty."""
        for col in range(DATA_START_COLUMN, DATA_END_COLUMN + 1):
            value = get_cell_value(sheet, row, col)
            if value:
                return False
        return True

    # --- XML Generation Logic ---
    
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
    except FileNotFoundError:
        error_list.append(f"FATAL ERROR: Excel file not found at path: {excel_path}")
        return None, error_list
    except Exception as e:
        error_list.append(f"FATAL ERROR: Could not load the Excel file. {e}")
        return None, error_list

    sheet = workbook.worksheets[0]
    
    # --- 1. Build Metadata ---
    root = Element('text')
    metadata = SubElement(root, 'text_metadata')
    for tag, cell_coord in METADATA_CELLS.items():
        cell = sheet[cell_coord]
        cell_value = get_cell_value(sheet, cell.row, cell.column)
        element = SubElement(metadata, tag)
        element.text = cell_value if cell_value else ""

    # --- 2. Process Interlinear Data (Body) ---
    body = SubElement(root, 'body')
    paragraph = SubElement(body, 'paragraph') 
    
    # CALCULATE TOTAL BLOCKS FOR PROGRESS BAR
    total_rows_to_scan = sheet.max_row - DATA_START_ROW + 1
    total_blocks = (total_rows_to_scan + ROWS_PER_LINE_BLOCK - 1) // ROWS_PER_LINE_BLOCK 
    
    # We use tqdm() to wrap the iteration for the progress bar
    for block_num in tqdm(range(total_blocks), desc="Processing Excel Blocks", unit="block"):
        
        current_row = DATA_START_ROW + (block_num * ROWS_PER_LINE_BLOCK)
        
        # Check if we've gone past the actual max row
        if current_row > sheet.max_row:
             break

        # Define the 4 rows for the current 'line' block
        vernacular_row = current_row
        gloss_row      = current_row + 1
        free_row       = current_row + 2
        blank_row      = current_row + 3 

        # --- Data Extraction ---
        vern_words = []
        gloss_words = []
        
        for col in range(DATA_START_COLUMN, DATA_END_COLUMN + 1):
            vern_val = get_cell_value(sheet, vernacular_row, col)
            gloss_val = get_cell_value(sheet, gloss_row, col)
            
            vern_is_present = bool(vern_val)
            gloss_is_present = bool(gloss_val)
            
            # CRITICAL CHECK: Alignment
            if vern_is_present != gloss_is_present:
                problem_cell = f"{chr(col + 64)}{vernacular_row}" if vern_is_present else f"{chr(col + 64)}{gloss_row}"
                error_list.append(
                    f"Alignment Error: Mismatched word/gloss at column {chr(col + 64)}. "
                    f"Non-empty cell: {problem_cell} (Rows {vernacular_row} and {gloss_row})."
                )
            
            if vern_is_present:
                vern_words.append(vern_val)
                gloss_words.append(gloss_val if gloss_val else "") 
            elif gloss_is_present:
                 pass # Ignore if only a gloss exists, but error is logged above
        
        free_translation = get_cell_value(sheet, free_row, DATA_START_COLUMN)
        is_block_empty = (not vern_words) and (not free_translation)

        if not is_block_empty:
            # --- Data Found: Process and Reset Counter ---
            consecutive_empty_blocks = 0 # Reset the counter

            line = SubElement(paragraph, 'line')
            il_lines = SubElement(line, 'il-lines')
            
            # 1. Vernacular Line
            vern_line = SubElement(il_lines, 'vernacular-line')
            for word in vern_words:
                wrd = SubElement(vern_line, 'wrd')
                wrd.text = word
                
            # 2. Gloss Line
            gloss_line = SubElement(il_lines, 'gloss-line')
            for gloss in gloss_words:
                gls = SubElement(gloss_line, 'gls')
                gls.text = gloss
            
            # 3. Free Translation
            free = SubElement(line, 'free')
            free.text = free_translation if free_translation else ""
            
            # Warning about the blank separator row
            if blank_row <= sheet.max_row and not is_row_empty(sheet, blank_row):
                 error_list.append(f"Warning: Expected blank separator row at Row {blank_row} is not empty.")

        elif is_block_empty and block_num > 0:
            # --- Paragraph Break / Early Exit Logic ---
            consecutive_empty_blocks += 1

            if consecutive_empty_blocks >= BLANK_BLOCK_EXIT_THRESHOLD:
                tqdm.write(f"\nExiting early: Found {BLANK_BLOCK_EXIT_THRESHOLD} consecutive empty blocks (approx. {BLANK_BLOCK_EXIT_THRESHOLD * ROWS_PER_LINE_BLOCK} blank rows).")
                break # Break the tqdm loop and exit processing
            
            # If it hasn't reached the threshold, signal a paragraph break if the current one has data
            if list(paragraph):
                paragraph = SubElement(body, 'paragraph')
            
    # Post-processing: Remove the last paragraph element if it ended up empty
    if not list(body) or not list(paragraph):
        for p in list(body):
            if not list(p):
                body.remove(p)
    
    return root, error_list


# ======================================================================
# --- CLI Execution Block (Isolated from conversion logic) ---
# ======================================================================

def prettify_xml(element):
    """Return a pretty-printed XML string for the given element."""
    # Ensure correct output encoding for the header
    rough_string = tostring(element, encoding='utf-8')
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent="  ")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert a specialized interlinear Excel spreadsheet into a UTF-8 XML file."
    )
    parser.add_argument(
        "input_file", 
        help="The path to the input Excel spreadsheet (.xlsx file)."
    )
    args = parser.parse_args()
    
    input_path = os.path.abspath(args.input_file)
    base_name, _ = os.path.splitext(input_path)
    output_xml_path = base_name + ".xml"
    error_log_path = base_name + "_processing_errors.txt"

    print(f"Starting conversion for: {os.path.basename(input_path)}")
    
    # 1. Run the core conversion function
    xml_root, errors = convert_excel_to_xml_dom(input_path)
    
    # Add an extra newline after the progress bar finishes to clean up the display
    print() 
    
    if xml_root is None:
        # Fatal error occurred (e.g., file not found, missing library)
        print("\n--- FATAL ERROR ---")
        for error in errors:
            print(error)
        sys.exit(1)
        
    # 2. Handle Errors
    if errors:
        with open(error_log_path, 'w', encoding='utf-8') as f:
            f.write(f"--- Processing Errors for {os.path.basename(input_path)} ---\n\n")
            for error in errors:
                f.write(f"- {error}\n")
        
        print(f"COMPLETED WITH WARNINGS ({len(errors)} found). See '{os.path.basename(error_log_path)}' for details.")
    else:
        if os.path.exists(error_log_path):
            os.remove(error_log_path)
        print("COMPLETED SUCCESSFULLY. No errors found.")
    
    # 3. Write XML Output
    try:
        pretty_xml = prettify_xml(xml_root)
        with open(output_xml_path, 'w', encoding='utf-8') as f:
            f.write(pretty_xml)
        print(f"XML output saved to: '{os.path.basename(output_xml_path)}'")
    except Exception as e:
        print(f"ERROR: Could not write XML file. {e}")
        sys.exit(1)