from abc import ABC, abstractmethod
import openpyxl
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom 


class InterlinearLoader(ABC):
    """
    Abstract class for loading an interlinear loader.

    Provides issuccess attribute and run() method.

    Concrete child classes must have:
    -a next_step attribute which defines a processing function
    """

    def __init__(self):
        self.issuccess = False
        super().__init__()  # for multiple inheritance...
        
        # a Loader object must define the following
        assert(hasattr(self, 'next_step'))
        assert(hasattr(self, '_progress'))

    @property
    def isdone(self):
        """
        Returns True if there are no more processing steps to take.
        """
        
        return bool(self.next_step is None)
    
    @property
    def progress(self):
        """
        Returns a float in the range of 0.0 - 1.0 representing processing progress.
        """

        return self._progress

    def run(self):
        """
        Run all steps directly (no breaks)
        """

        while not self.isdone:
            self.next_step()


class InterlinearXML:
    """
    An XML object for interlinear data.

    The XML root contains a "metadata" element and a "body" element.
    New paragraphs are added to the body.
    New line are added to the current paragraph.
    Similarly, all other new elements are added to the current parent element.
    There is no class-specific approach to retrieving previous lines, paragraphs, etc.
      but the XML objects have methods for this.

    InterlinearXML Methods:
      (new_xml_* take no arguments)
        new_xml_paragraph()
        new_xml_line()
        new_xml_il_lines()
        new_xml_vernacular_line()
        new_xml_gloss_line()
      (add_xml_* take text argument)
        add_xml_vernacular_word(text)
        add_xml_gloss_word(text)
        add_xml_free(text)
      (output handling)
        get_pretty_xml()
        write(filename)
    """

    def __init__(self):
        # does it work fine to define body before filling metadata?
        self.xml_root = Element('text')
        self.xml_metadata = SubElement(self.xml_root, 'text_metadata')
        self.xml_body = SubElement(self.xml_root, 'body')
        self.new_xml_paragraph()

        super().__init__()  # for multiple inheritance...
    
    def new_xml_paragraph(self):
        self.xml_paragraph = SubElement(self.xml_body, 'paragraph')

    def new_xml_line(self):
        self.xml_line = SubElement(self.xml_paragraph, 'line')

    def new_xml_il_lines(self):
        self.xml_il_lines = SubElement(self.xml_line, 'il-lines')

    def new_xml_vernacular_line(self):
        self.xml_vern_line = SubElement(self.xml_il_lines, 'vernacular-line')

    def new_xml_gloss_line(self):
        self.xml_gloss_line = SubElement(self.xml_il_lines, 'gloss-line')

    def add_xml_vernacular_word(self, text):
        wrd = SubElement(self.xml_vern_line, 'wrd')
        wrd.text = text

    def add_xml_gloss_word(self, text):
        wrd = SubElement(self.xml_gloss_line, 'gls')
        wrd.text = text

    def add_xml_free(self, text):
        xml_free = SubElement(self.xml_line, 'free')
        xml_free.text = text

    def get_pretty_xml(self):
        """
        Return a pretty-printed XML string for the root XML.
        """

        # Ensure correct output encoding for the header
        rough_string = tostring(self.xml_root, encoding='utf-8')
        reparsed = minidom.parseString(rough_string)
        return reparsed.toprettyxml(indent="  ")
    
    def write(self, filename):
        """
        Write pretty-printed XML to file.
        """

        pretty = self.get_pretty_xml()
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(pretty)


class ExcelInterlinearLoader(InterlinearLoader, InterlinearXML):
    """
    Handles loading of an Excel interlinear from template.

    Direct usage:
        e = ExcelInterlinearLoader(name_and_path_of_excel_file_to_load)
        e.run()
        txt = e.get_pretty_xml()

    To use tqdm for displaying progress:
        e = ExcelInterlinearLoader(name_and_path_of_excel_file_to_load)
        with tqdm(total=1.0, desc="Processing Excel File") as pbar:
            while not e.isdone:
                e.next_step()
                pbar.update(e.progress)
    """

    def __init__(self, loadname):
        self.METADATA_CELLS = {
            'title': 'C2',
            'author': 'C3',
            'transcriber': 'C4',
            'writing_system_vernacular': 'N2',
            'writing_system_free': 'N3',
            'writing_system_gloss': 'N4'
        }
        self.DATA_START_ROW = 6
        self.DATA_START_COLUMN = 3 # Column C
        self.DATA_END_COLUMN = 26 # Column Z (where the free translation merge ends)
        self.ROWS_PER_LINE_BLOCK = 4
        self.BLANK_BLOCK_EXIT_THRESHOLD = 5 # Exit after 5 consecutive empty 4-row blocks (20 blank rows)
        self.FILE_LOAD_PROGRESS_WEIGHT = 0.5

        self.warning_list = []  # fatal errors are raised as exceptions to be handled elsewhere
        self.consecutive_empty_blocks = 0

        self.loadname = loadname
        self.n_blocks = None
        self._progress = 0
        self.current_block = None
        self.next_step = self.load_sheet
        super().__init__()

        self.debug = False  # for printing to console

    def update_progress(self, value=None):
        """
        Update the progress attribute, either from current_block or with a set value.

        load_sheet is assigned self.FILE_LOAD_PROGRESS_WEIGHT of the progress.
        The rest of the progressbar is evenly divided among:
            read_metadata: 1 unit 
            read_one_block: 1 unit per block
            cleanup: 1 unit
        """

        w = self.FILE_LOAD_PROGRESS_WEIGHT

        if value is None:
            self._progress = w + (1 - w) * self.current_block / (self.n_blocks + 2)
        elif value == -1:
            self._progress = w + (1 - w) * (self.n_blocks + 1) / (self.n_blocks + 2)
        else:
            self._progress = float(value)

    def load_sheet(self):
        """
        Load the Excel sheet as an openpyxl object and count the rows.
        """

        try:
            workbook = openpyxl.load_workbook(self.loadname, data_only=True)
        except Exception as e:
            raise Exception(f"Error loading Excel file '{self.loadname}'") from e

        try:
            self.sheet = workbook.worksheets[0]
        except Exception as e:
            raise Exception(f"Error loading first sheet of Excel file '{self.loadname}'") from e

        n_data_rows = self.sheet.max_row - self.DATA_START_ROW + 1
        if self.debug:
            print(f'  Total rows: {self.sheet.max_row}')
            print(f'  Data rows: {n_data_rows}')
        if n_data_rows % self.ROWS_PER_LINE_BLOCK:
            # extra row or part of an interlinear line
            n_extra = n_data_rows % self.ROWS_PER_LINE_BLOCK
            self.warning_list.append('Partial interlinear block of data will be ignored ' +
                                     f'({n_extra} extra data rows found)')
        self.n_blocks = (n_data_rows // self.ROWS_PER_LINE_BLOCK)
        self.update_progress(self.FILE_LOAD_PROGRESS_WEIGHT)
        self.next_step = self.read_metadata
        # need approach for if a step fails, how to let GUI know?
        # self.next_step = None
        # self.isdone = True
        # self.success = False    # something like that
        if self.debug:
            print(f'load_sheet: n_blocks = {self.n_blocks}')

    def read_metadata(self):
        """
        Read the metadata cells of the spreadsheet
        """
        
        for tag, cell_coord in self.METADATA_CELLS.items():
            cell = self.sheet[cell_coord]
            cell_value = self.get_cell_value(cell.row, cell.column)
            element = SubElement(self.xml_metadata, tag)
            element.text = cell_value if cell_value else ""
        self.current_block = 1
        self.update_progress()
        self.next_step = self.read_one_block
    
    def read_one_block(self):
        """
        Read one interlinear line (block) of data, and check if done.
        """
        
        vernacular_row = self.DATA_START_ROW + (self.current_block - 1) * self.ROWS_PER_LINE_BLOCK
        gloss_row =      vernacular_row + 1
        free_row =       vernacular_row + 2
        # blank_row =      vernacular_row + 3 # worth checking for blankness or no?

        vern_words = []
        gloss_words = []

        for col in range(self.DATA_START_COLUMN, self.DATA_END_COLUMN+1):
            vern_val = self.get_cell_value(vernacular_row, col)
            gloss_val = self.get_cell_value(gloss_row, col)

            vern_is_present = bool(vern_val)
            gloss_is_present = bool(gloss_val)

            # Check alignment
            if vern_is_present != gloss_is_present:
                problem_cell = f"{chr(col + 64)}{vernacular_row}" if vern_is_present else f"{chr(col + 64)}{gloss_row}"
                self.warning_list.append(
                    f"Alignment Error: Mismatched word/gloss at column {chr(col + 64)}. "
                    f"Non-empty cell: {problem_cell} (Rows {vernacular_row} and {gloss_row})."
                )
            if vern_is_present:
                vern_words.append(vern_val)
                gloss_words.append(gloss_val if gloss_val else "")
            elif gloss_is_present:
                pass # Ignore if only a gloss exists, but error is logged above
        free_translation = self.get_cell_value(free_row, self.DATA_START_COLUMN)
        if free_translation is None:
            free_translation = ""
        is_block_empty = (not vern_words) and (not free_translation)

        if not is_block_empty:
            self.consecutive_empty_blocks = 0
            self.new_xml_line()
            self.new_xml_il_lines()
            self.new_xml_vernacular_line()
            for word in vern_words:
                self.add_xml_vernacular_word(word)
            self.new_xml_gloss_line()
            for word in gloss_words:
                self.add_xml_gloss_word(word)
            self.add_xml_free(free_translation)
        # elif is_block_empty and self.current_row > self.DATA_START_ROW:
        # ^^^ why the 2nd condition? why ignore the first blank line?
        else:
            # Paragraph break / Early Exit Logic
            self.consecutive_empty_blocks += 1
            if self.consecutive_empty_blocks >= self.BLANK_BLOCK_EXIT_THRESHOLD:
                # self.warning_list.append(
                #     f"Finishing early due to {self.consecutive_empty_blocks} consecutive empty interlinear lines")
                self.update_progress(-1)
                self.next_step = self.cleanup
                return None
            elif list(self.xml_paragraph):
                self.new_xml_paragraph()    

        self.current_block += 1
        self.update_progress()
        if self.current_block > self.n_blocks:
            # self.update_progress(-1)
            self.next_step = self.cleanup

    def cleanup(self):
        """
        Post-processing: Remove the last paragraph element if it ended up empty.

        Then indicate that the processing is completed.
        """

        if not list(self.xml_body) or not list(self.xml_paragraph):
            for p in list(self.xml_body):
                if not list(p):
                    self.xml_body.remove(p)
        self.next_step = None
        self.update_progress(1.0)
        self.issuccess = True

    def get_cell_value(self, row, col):
        """
        Get value of one cell of self.sheet, cleanly
        """

        cell = self.sheet.cell(row=row, column=col)
        if cell.value is None:
            return None
        else:
            return str(cell.value).strip()


if __name__ == "__main__":
    # TEMP: for testing
    import os
    # filename = r'Cerita Juari Atau (Barnabas) - in template.xlsx'
    filename = r'Interlinear Text Excel Template (80 lines)2.xlsx'
    xl = ExcelInterlinearLoader(filename)

    if False:
        from tqdm import tqdm
        with tqdm(total=1.0, desc="Processing Excel File") as pbar:
            while xl.next_step is not None:
                xl.next_step()
                pbar.update(xl.progress)
        outputname = filename[:-5] + r'_ClassTestTqdm.xml'
        xl.write(outputname)
        print(f'Output written to: {outputname}')
        print('')
        print(f'issuccess: {xl.issuccess}')
        print(f'next_step: {xl.next_step}')
        print(f'warning_list: {xl.warning_list}')
    else:
        xl.run()
        outputname = filename[:-5] + r'_ClassTest.xml'
        xl.write(outputname)
        print(f'Output written to: {outputname}')
        print('')
        print(f'issuccess: {xl.issuccess}')
        print(f'next_step: {xl.next_step}')
        print(f'warning_list: {xl.warning_list}')